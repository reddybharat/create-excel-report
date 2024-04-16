"""
    How to create an executable file?
    Steps :
    1. Open Terminal and navigate to location where script is located (using cd)
       OR
       Open terminal wherever the script is located directly
    
    If pyinstaller is not installed, type pip install pyinstaller  
    2. pyinstaller --onefile file_name.py

    Note: Put the input file in the same folder as exe file (dist folder) before running the exe
"""


#imports
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime


exe_path = os.path.dirname(sys.executable)

input_path = os.path.join(exe_path, 'beverage_sales_data.xlsx')

#create pivot table
df = pd.read_excel(input_path)
df = df[['country','sales','brand']]

df_brands = df['brand'].unique()
print(df_brands)
brand_input = input('\nType brand name here (same as shown above): ')

df = df[df['brand'] == brand_input]
pivot_table = df.pivot_table(index='brand', columns='country', values='sales', aggfunc=sum)
out_pivot_path = os.path.join(exe_path, 'temp_pivot_table.xlsx')
pivot_table.to_excel(out_pivot_path, 'Report', startrow=4)

#creating the excel report
wb = load_workbook(out_pivot_path) #reading the same pivot table we saved
sheet = wb['Report']

#creating barchart
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B9")
barchart.title = f"{brand_input} Sales"
barchart.style = 2

for i in range(min_row+1, max_row+1):
    # letter = get_column_letter(i)
    sheet[f'O{i}'] = f'=SUM(B{i}:N{i})'
    # print(f'=SUM({letter}{min_column+1}:{letter}{max_column})')
    sheet[f'O{i}'].style = 'Currency'

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row}'].style = 'Currency'

sheet['A1'] = "Sales By Country"
sheet['A2'] = brand_input
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=15)
sheet['O5'] = 'TOTAL'
sheet['O5'].font = Font('Arial', bold=True)


#saving the final excel report
now = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
output_path = os.path.join(exe_path, f'{brand_input}_Report_{now}.xlsx')
wb.save(output_path)
os.remove(out_pivot_path) #delete the temporary pivot table
